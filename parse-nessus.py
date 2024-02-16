#!/usr/bin/env python3

import argparse
import xml.etree.ElementTree as ET
import os
from datetime import datetime
import openpyxl


csvHeaders = ['Severity', 'CVSS-v3 score', 'IP Address', 'Ports', 'Vulnerability', 'CVE\'s']
nessusFields = ['risk_factor', 'cvss3_base_score', 'host-ip', 'port', 'plugin_name', 'cve']
reportRows = []
findings = []


# Clean values from Nessus report
def getValue(rawValue):
    if rawValue == None:
        rawValue = "empty"
    else:
        cleanValue = rawValue.replace('\n', ' ').strip(' ')
        if len(cleanValue) > 32000:
            cleanValue = cleanValue[:32000] + ' [Trimmed due to length]'
        return cleanValue


# Helper function for handleReport()
def getKey(rawKey):
    return csvHeaders[nessusFields.index(rawKey)]


# Handle a single report item
def handleReport(report):
    findings = []
    updated_data = []

    reportHost = dict.fromkeys(csvHeaders, '')
    for item in report:
        if item.tag == 'HostProperties':
            for tag in (tag for tag in item if tag.attrib['name'] in nessusFields):
                reportHost[getKey(tag.attrib['name'])] = getValue(tag.text)
        if item.tag == 'ReportItem':
            reportRow = dict(reportHost)
            reportRow['Ports'] = item.attrib['port']
            reportRow['Vulnerability'] = item.attrib['pluginName']
            reportRow['pluginId'] = item.attrib['pluginID']
            cvelist = []

            for tag in (tag for tag in item if tag.tag in nessusFields):
                key = getKey(tag.tag)
                value = getValue(tag.text)
                
                if key == 'Severity':
                    if value == 'Critical':
                        value = '0. Critical'
                    elif value == 'High':
                        value = '1. High'
                    elif value == 'Medium':
                        value = '2. Medium'
                    elif value == 'Low':
                        value = '3. Low'
                    elif value == 'None':
                        value = '4. Info'

                if key == 'CVE\'s':
                    cvelist.append(value)

                reportRow[key] = value

            reportRow['CVE\'s'] = ', '.join(cvelist)

            findings.append(reportRow)

    for entry in findings:
        cvss_score = entry['CVSS-v3 score']
        sev = entry['Severity']
        severity = map_cvss_to_severity(cvss_score, sev)
        entry['Severity'] = severity
        updated_data.append(entry)
    return findings


# Function to map CVSS-v3 score to severity
def map_cvss_to_severity(cvss_score,sev):
    cvss_score.replace('*','')
    if cvss_score == '':
        return sev
    elif float(cvss_score) >= 9.0:
        return '0. Critical'
    elif float(cvss_score) >= 7.0:
        return '1. High'
    elif float(cvss_score) >= 4.0:
        return '2. Medium'
    else:
        return '3. Low'

# Get files 
def getargs():
    parser = argparse.ArgumentParser(description="Merge all .nessus files within a folder into one .csv report in that folder", formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(type=str, dest="directory", help="Folder containing .nessus files")
    parser.add_argument('--split', action='store_true', help="Split data")
    parser.add_argument('--minsev', type=str, help="Severity Filter (critical, high, medium, low)")
    args = parser.parse_args()
    return args


def process_nessus_files():
    unique_report_rows = []
    unique_report_rows_merged = []

    merged_rows = {}
    unique_lines = set()

    for file in nessusFiles:
        tree = ET.parse(file)
        root = tree.getroot()
        try:
            scanFile = ET.parse(file)
            xmlRoot = scanFile.getroot()
            for report in xmlRoot.findall('./Report/ReportHost'):
                rootReport = root.find('Report')
                for report in xmlRoot.findall('./Report/ReportHost'):
                    findings = handleReport(report)###
                    reportRows.extend(findings)
        except IOError:
            print("Could not find file \"" + file + "\"")

    try:
        for D in reportRows:
            line_tuple = tuple(D.items())
            if line_tuple not in unique_lines:

                unique_lines.add(line_tuple)            
                unique_report_rows.append(D)
    except Exception as e:
        print("An error occurred while processing rows:", e)

    try:

        for D in unique_report_rows:
            line_tuple = tuple(D.items())
            ip_address = D['IP Address']
            plugin_id = D['pluginId']
            port = D['Ports']
            
            # Key for grouping
            key = (ip_address, plugin_id)
            
            # Check if the key already exists
            if key in merged_rows:
                # If yes, update the port value by concatenating the new port
                merged_rows[key]['Ports'] += f", {port}"
            else:
                # If not, add a new entry
                merged_rows[key] = D

        unique_report_rows_merged = list(merged_rows.values())

    except Exception as e:
        print(f"An error occurred: {e}")

    return sorted(unique_report_rows_merged, key=lambda x: (x['Severity'], -float(x.get('CVSS-v3 score', 0) or 0)))


def createExcelFile():
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    return os.path.join(f'{args.directory}_{timestamp}.xlsx')


def createWb():
    wb = openpyxl.Workbook()
    return wb


def fillempty(ws, sorted_report_rows):
    for D in sorted_report_rows:
        row_values = []
        for header in csvHeaders:
            value = D.get(header, '')
            if value == '':
                value = '*'
            row_values.append(value)
        ws.append(row_values)


def printonetable(ws, sorted_report_rows):
    table = openpyxl.worksheet.table.Table(displayName="Table1", ref=ws.dimensions)
    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(csvHeaders)):
        textstyle(row)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=len(csvHeaders))):
        rowstyle(row_idx, row)
        
    headerstyle(ws[1])
    severitystyle(ws)


def printmultipletables(ws, sorted_report_rows):
    ip_tables = {}

    for row in sorted_report_rows:
        ip = row.get('IP Address')
        if ip:
            if ip not in ip_tables:
                ip_tables[ip] = []
            ip_tables[ip].append(row)

    totallength = 2

    for ip, rows in ip_tables.items():
        ws.append([f"IP Address: {ip}"])
        ws.append(csvHeaders)
        fillempty(ws, rows)
        ws.append([]) 

        start_row = totallength
        end_row = start_row + len(rows)
        totallength = 3 + end_row
        ref = f"A{start_row}:{openpyxl.utils.get_column_letter(len(csvHeaders))}{end_row}"
        table = openpyxl.worksheet.table.Table(displayName=f"Table_{ip.replace('.', '_')}", ref=ref)
        style = openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
    
        for row in ws.iter_rows(min_row=start_row-1, max_row=end_row, min_col=1, max_col=len(csvHeaders)):
            textstyle(row)

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row+1, max_row=end_row, min_col=2, max_col=len(csvHeaders))):
            rowstyle(row_idx, row)

        headerstyle(ws[start_row])
        severitystyle(ws)


def headerstyle(header_row):
    for cell in header_row:
        cell.fill = openpyxl.styles.PatternFill(start_color="F05025", end_color="F05025", fill_type="solid")
        cell.font = openpyxl.styles.Font(name="Century Gothic", bold=True, color="FFFFFF", size=10) 
        border = openpyxl.styles.Side(style="thin", color="F05025")
        cell.border = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)


def severitystyle(ws):
    for row_idx, D in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1), start=2):
        severity_value = D[0].value
        cell = D[0] 
        
        if severity_value == '0. Critical':
            cell.value = 'Critical'
            fill_color = "C00000"
        elif severity_value == '1. High':
            cell.value = 'High'
            fill_color = "FF0000" 
        elif severity_value == '2. Medium':
            cell.value = 'Medium'
            fill_color = "FFC000" 
        elif severity_value == '3. Low':
            cell.value = 'Low'
            fill_color = "FFFF00" 
        elif severity_value == '4. Info':
            cell.value = 'Info'
            fill_color = "00B0F0" 
        else:
            fill_color = None
        
        if fill_color:
            fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            border = openpyxl.styles.Side(style="thin", color="F05025")
            cell.border = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)
            cell.fill = fill


def textstyle(row):
    for cell in row:
        cell.font = openpyxl.styles.Font(name="Century Gothic", size=10)


def rowstyle(row_idx, row):
    for col_idx, cell in enumerate(row):
        border = openpyxl.styles.Side(style="thin", color="F05025")
        cell.border = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)
        fill_color = "FCDBD3" if row_idx % 2 == 0 else "FFFFFF"
        fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.fill = fill


def filterreportrows(arg, sorted_report_rows):
    if arg.lower() == 'critical':
        return [row for row in sorted_report_rows if not row['Severity'].startswith(('1', '2', '3', '4'))]
    elif arg.lower() == 'high':
        return [row for row in sorted_report_rows if not row['Severity'].startswith(('2', '3', '4'))]
    elif arg.lower() == 'medium':
        return [row for row in sorted_report_rows if not row['Severity'].startswith(('3', '4'))]
    elif arg.lower() == 'low':
        return [row for row in sorted_report_rows if not row['Severity'].startswith(('4'))]
    else:
        print('[!] Cannot interpret value for --minsev')
        exit()


# Main
if __name__ == '__main__':
    args = getargs()
    if not os.path.isdir(args.directory):
        print('[!] Cannot find specified directory')
        exit()

    # find all .nessus files in the directory
    nessusFiles = [os.path.join(args.directory, file) for file in os.listdir(args.directory) if file.endswith('.nessus')]

    if len(nessusFiles) == 0:
        print('[!] No .nessus files found!')
        exit()
    else:
         print(f'[*] Found {len(nessusFiles)} nessus files!')

    sorted_report_rows = process_nessus_files()
    
    if args.minsev:
        sorted_report_rows = filterreportrows(args.minsev, sorted_report_rows)

    excelfile = createExcelFile()
    wb = createWb()
    ws = wb.active

    if args.split:
        printmultipletables(ws, sorted_report_rows)
    else:    
        ws.append(csvHeaders)
        fillempty(ws, sorted_report_rows)
        printonetable(ws, sorted_report_rows)

    wb.save(excelfile)
